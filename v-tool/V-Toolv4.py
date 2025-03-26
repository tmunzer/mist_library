#!/usr/bin/env python3
# ------------------------------------------------------------------------------
# MIT License
#
# Copyright (c) 2025 Steve Voto (Svoto)
#
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"), to deal
# in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
# copies of the Software, and to permit persons to whom the Software is furnished
# to do so, subject to the following conditions:
#
# The above copyright notice and this permission notice shall be included in all
# copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
# SOFTWARE.
#
# ------------------------------------------------------------------------------
#
# Liability Release:
# This code is provided for testing purposes only. Use at your own risk.
# The author assumes no liability for any damages or losses incurred
# through the use or misuse of this script.
# ------------------------------------------------------------------------------
"""
===============================================================================
Upgrade Management and Network Operations Script
===============================================================================
Version: 4.0 Full Version

Description:
------------
This tool is an experimental utility designed for testing and educational purposes only.
It should not be used in production environments. 
Built to demonstrate the capabilities of modern MIST APIs, this tool streamlines network 
device management and upgrade workflows by consolidating key functions such as inventory retrieval, 
device upgrade initiation, live status monitoring, and comprehensive report generation.

1. **Inventory Retrieval**:
   - Fetches gateway inventory data from the Mist API.
   - Saves the inventory into a formatted Excel file with headers and auto-adjusted columns.

2. **Gateway Upgrade Actions**:
   - Processes upgrade commands based on an Excel inventory.
   - Initiates upgrade requests to gateways via the Mist API.

3. **Live Upgrade Monitoring**:
   - Displays real-time upgrade status with color-coded output.
   - Supports non-blocking key detection to allow early exit back to the main menu.

4. **Reporting**:
   - Generates various reports including:
       - Physical Device Interface Report.
       - Full Device Interface Statistics Report.
       - Device Peer Path Status Report.
   - Access Point Status Report: Retrieves AP status for each site and merges router names from inventory.
   - Post Reports: Creates consolidated reports for devices marked for upgrade (prefixed with "Post-").

5. **Backup and Restore Operations**:
   - **Backup Org**: Collects configuration data (including site settings, network configurations, and service policies) 
     from the Mist API, saving each as JSON files in a timestamped backup folder.
   - **Full Site Restore**: Allows selection of a backup folder to restore site configurations 
     and related settings, updating organization IDs as necessary.
   
6. **Organization Cleanup**:
   - Provides an interactive deletion tool to list and remove various organizational resources 
     (Sites, Applications, Networks, Hub Profiles, WAN Edges, and Switches).

7. **Exit**:
   - Allows the user to exit the program safely.

===============================================================================
"""

# ---------------------------
# Imports
# ---------------------------
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
import inspect  # To get function name dynamically
import os
import sys
import json
import time
import select
import re
import requests
import argparse
import pandas as pd
from datetime import datetime
import termios
import tty


# ---------------------------
# Display Menu with Word Art
# ---------------------------
def display_menu():
    """Displays the ASCII word art menu."""
    print("""
*****************************************************************
* █▓▒▒█▓▒▒░░░█▓▒▒░░░   V-Tool Version 4.0  ░░░▒▒▓█░░░▒▒▓█▓█▒▒▓█ *
*****************************************************************
* █▓ Press "x" anytime to abort and return to previous menu ▓█  *
*****************************************************************
""")
    print("\nMenu:")
    print("1. View Inventory")
    print("2. Run Upgrades")
    print("3. Check Upgrade Status")
    print("4. Physical Device Interface Report")
    print("5. Full Device Interface Stats Report")
    print("6. Device Peer Path Status Report")
    print("7. Access Point Status")
    print("8. Post Reports")
    print("9. Backup Org")
    print("10. Restore Org Items")
    print("11. Remove Org Items")
    print("0. Exit")

# ---------------------------
# Helper Function: Replace org_id
# ---------------------------
def replace_org_id(data, new_org_id):
    """
    Recursively replaces any key named 'org_id' in the JSON object or list with new_org_id.
    """
    if isinstance(data, dict):
        for key, value in data.items():
            if key.lower() == "org_id":
                data[key] = new_org_id
            elif isinstance(value, (dict, list)):
                data[key] = replace_org_id(value, new_org_id)
    elif isinstance(data, list):
        data = [replace_org_id(item, new_org_id) for item in data]
    return data


# ---------------------------
# Read API Credentials
# ---------------------------
def read_token_org_url(file_path="Token-Org-URL.txt"):
    """Reads API token, org_id, and base_url from a file."""
    try:
        with open(file_path, 'r') as f:
            credentials = {line.split("=")[0].strip(): line.split("=")[1].strip() for line in f if "=" in line}
        return credentials["token"], credentials["org_id"], credentials["base_url"]
    except FileNotFoundError:
        print(f"[ERROR] File {file_path} not found.")
        sys.exit(1)
    except KeyError as e:
        print(f"[ERROR] Missing {e} in {file_path}.")
        sys.exit(1)

# ---------------------------
# Convert Epoch Time to Readable Format
# ---------------------------
def convert_epoch_to_datetime(epoch_time):
    """Converts epoch time to human-readable datetime."""
    return datetime.utcfromtimestamp(epoch_time).strftime('%Y-%m-%d %H:%M:%S') if epoch_time else "N/A"

# ---------------------------
# Find Latest Inventory File
# ---------------------------
def get_latest_inventory_file():
    """Finds the newest inventory file with prefix "inventoryStats" (X being the highest number)."""
    files = [f for f in os.listdir() if f.startswith("inventoryStats") and f.endswith(".xlsx")]
    if not files:
        return None
    def extract_num(filename):
        m = re.search(r'inventoryStats(\d+)\.xlsx', filename)
        return int(m.group(1)) if m else 0
    return sorted(files, key=lambda x: extract_num(x))[-1]
# ---------------------------
# Fetch Inventory from Mist API (Menu Item 1)
# ---------------------------
def get_gateway_inventory(org_id, token, base_url):
    """Fetches gateway inventory from Mist API."""
    url = f"{base_url}/orgs/{org_id}/inventory?type=gateway"
    headers = {
        "Authorization": f"Token {token}",
        "Content-Type": "application/json",
        "Accept": "application/json"
    }
    print(f"[INFO] Fetching gateway inventory from: {url}")
    response = requests.get(url, headers=headers)
    log_api_request("GET", "get_gateway_inventory", url, None, response)
    if response.status_code != 200:
        print(f"[ERROR] GET request failed. HTTP {response.status_code}: {response.text}")
        log_error("GET request failed in Menu Item 1 (View Inventory)", extra_info=f"HTTP {response.status_code}: {response.text}")
        return None
    try:
        data = response.json()
        print(f"[INFO] Retrieved {len(data)} gateways.")
        return data
    except json.JSONDecodeError:
        print("[ERROR] Failed to parse JSON response.")
        log_error("JSON parse error in Menu Item 1 (View Inventory)")
        return None

# ---------------------------
# Save Data to Excel with Formatting
# ---------------------------
# ---------------------------
# Save Data to Excel with Auto-Increment & Ensure Upgrade Columns
# ---------------------------
def save_to_excel(data, filename=None):
    """
    Saves data to an Excel file, auto-increments filename if necessary,
    and ensures missing upgrade-related columns are included.
    """
    base_filename = "inventoryStats"
    ext = ".xlsx"

    # Find the next available filename if not provided
    if filename is None:
        latest_num = 1
        while os.path.exists(f"{base_filename}{latest_num}{ext}"):
            latest_num += 1
        filename = f"{base_filename}{latest_num}{ext}"

    elif os.path.exists(filename):  # If provided filename exists, increment its number
        base, ext = os.path.splitext(filename)
        i = 1
        while os.path.exists(f"{base}{i}{ext}"):
            i += 1
        filename = f"{base}{i}{ext}"

    # Convert data to DataFrame if not already
    df = data if isinstance(data, pd.DataFrame) else pd.DataFrame(data)

    # Ensure missing columns are present
    required_columns = ["Action", "Upgrade Version", "Upgrade Status", "Upgrade Progress"]
    for col in required_columns:
        if col not in df.columns:
            df[col] = ""  # Default to empty string

    # Save to Excel with formatting
    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        df.to_excel(writer, index=False)
        wb = writer.book
        worksheet = wb.active

        # Formatting: Hunter Green Header
        header_fill = PatternFill(start_color="355E3B", end_color="355E3B", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font

        # Auto-adjust column widths
        for column_cells in worksheet.columns:
            max_length = 0
            column = column_cells[0].column_letter
            for cell in column_cells:
                if cell.value:
                    cell_value = str(cell.value).replace("\n", " ")
                    max_length = max(max_length, len(cell_value))
            worksheet.column_dimensions[column].width = max_length + 2

    print(f"[INFO] Data saved to {filename}")

# ---------------------------
# Log Error Handling
# ---------------------------
def log_error(message, extra_info=""):
    """Appends an error message with a timestamp and extra details to the current log file."""
    log_file = "log.xlsx"

    if not os.path.exists(log_file):
        wb = Workbook()
        ws = wb.active
        ws.append(["Timestamp", "Error Message", "Details"])
        wb.save(log_file)

    try:
        wb = load_workbook(log_file)
        ws = wb.active
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws.append([timestamp, message, extra_info])
        wb.save(log_file)
    except Exception as e:
        print(f"[LOGGING ERROR] Could not write to log file: {e}")

# ---------------------------
# Perform Upgrade (Menu Item 2)
# ---------------------------
def perform_upgrade(mac, site_id, version, token, base_url):
    """Performs a POST request to upgrade the gateway using the working API endpoint."""
    upgrade_url = f"{base_url}/sites/{site_id}/ssr/00000000-0000-0000-1000-{mac}/upgrade"
    upgrade_data = {
        "channel": "stable",
        "version": version,
        "start_time": None,
        "reboot_at": None
    }
    print(f"[INFO] Sending upgrade request for {mac} to version {version}.")
    response = requests.post(
        upgrade_url,
        headers={"Authorization": f"Token {token}", "Content-Type": "application/json"},
        data=json.dumps(upgrade_data)
    )
    if response.status_code == 200:
        print(f"[SUCCESS] Upgrade request sent for {mac} to version {version}.")
        return True
    else:
        print(f"[ERROR] Upgrade failed for {mac}. HTTP {response.status_code}: {response.text}")
        log_error("Upgrade failed in Run Upgrades", extra_info=f"{mac} - HTTP {response.status_code}: {response.text}")
        return False

# ---------------------------
# Process Upgrades (Menu Item 2)
# ---------------------------
def process_inventory_actions(filename, token, base_url):
    """Processes upgrade actions from the inventory file."""
    
    # Read the Excel file using openpyxl
    df = pd.read_excel(filename, engine="openpyxl")
    
    # Ensure necessary columns exist and set data type to string to prevent float conversion errors
    if "Upgrade Status" not in df.columns:
        df["Upgrade Status"] = ""
    
    df["Upgrade Status"] = df["Upgrade Status"].astype(str)  # Convert to string

    for index, row in df.iterrows():
        if str(row.get("Action", "")).strip().upper() == "U":
            site_id = row.get("site_id", "")
            mac = row.get("mac", "")
            version = str(row.get("Upgrade Version", "")).strip()

            if site_id and mac and version:  # Ensure valid values
                success = perform_upgrade(mac, site_id, version, token, base_url)

                if success:
                    df.at[index, "Upgrade Status"] = "Upgrade Requested"
    
    # Save the updated DataFrame back to the Excel file
    df.to_excel(filename, index=False, engine="openpyxl")

# ---------------------------
# Check Upgrade Status (Menu Item 3)
# ---------------------------
def check_upgrade_status(filename, token, base_url):
    """
    Checks and displays the upgrade status for devices marked for upgrade.
    This version uses each device's site_id and mac to build the status URL and extracts the "fwupdate" data.
    """
    print("\n[INFO] Press 'X' to return to the main menu.")
    
    df = pd.read_excel(filename, engine="openpyxl")
    
    # Ensure necessary columns exist
    if "Upgrade Version" not in df.columns:
        df["Upgrade Version"] = ""
    if "Upgrade Status" not in df.columns:
        df["Upgrade Status"] = ""
    if "Upgrade Progress" not in df.columns:
        df["Upgrade Progress"] = 0  # Default progress to 0

    # Ensure "Upgrade Status" is treated as a string
    df["Upgrade Status"] = df["Upgrade Status"].astype(str)

    devices_to_check = df[df["Action"].astype(str).str.strip().str.upper() == "U"]
    if devices_to_check.empty:
        print("[INFO] No devices marked for upgrade.")
        return

    while True:
        os.system("clear")
        print(f"\033[40m\033[97m{'Name':<15} {'MAC':<15} {'Status':<15} {'Progress':<10} {'Version':<10}\033[0m")
        print("-" * 70)

        for index, row in devices_to_check.iterrows():
            site_id = row.get("site_id", "")
            mac = row.get("mac", "")
            name = row.get("name", "Unknown")

            if not site_id or not mac:
                continue  # Skip rows with missing data

            status_url = f"{base_url}/sites/{site_id}/stats/devices/00000000-0000-0000-1000-{mac}?type=gateway"

            try:
                response = requests.get(status_url, headers={"Authorization": f"Token {token}"})
                log_api_request("GET", "check_upgrade_status", status_url, None, response)

                if response.status_code == 200:
                    data = response.json()
                    fwupdate = data.get("fwupdate", {})
                    status = fwupdate.get("status", "IDLE")
                    progress = fwupdate.get("progress", 0)
                    version = data.get("version", None)

                    color_map = {
                        "downloading": "\033[33m",
                        "rebooting": "\033[31m",
                        "upgrading": "\033[35m",
                        "connected": "\033[32m"
                    }
                    color = color_map.get(status.lower(), "\033[0m") if status else "\033[0m"

                    print(f"{color}{name:<15} {mac:<15} {str(status).upper():<15} {str(progress)+'%':<10} {str(version or 'N/A'):<10}\033[0m")
                    
                    # Ensure values are stored correctly in DataFrame
                    df.at[index, "Upgrade Status"] = str(status)
                    df.at[index, "Upgrade Progress"] = float(progress) if isinstance(progress, (int, float)) else 0
                else:
                    print(f"[ERROR] HTTP error {response.status_code} for site_id {site_id}")
                    log_error("HTTP error in Check Upgrade Status", extra_info=f"Site_id {site_id} - HTTP {response.status_code}")

            except Exception as e:
                print(f"[ERROR] Exception for site_id {site_id}: {e}")
                log_error("Exception in Check Upgrade Status", extra_info=str(e))
                continue

            time.sleep(1)

        # Save the updated DataFrame
        df.to_excel(filename, index=False, engine="openpyxl")

        print("-" * 70)
        print("[INFO] Refreshing in 30 seconds. Press 'X' to return to the main menu immediately.")

        exit_flag = False
        for i in range(30):
            key = wait_for_keypress(1)
            if key and key.strip().lower() == "x":
                print("\n[INFO] Returning to main menu.")
                exit_flag = True
                break
        if exit_flag:
            break

# ---------------------------
# Retrieve Physical Device Interface Report (Menu Item 4)
# ---------------------------
def retrieve_physical_device_interface_report(token, org_id, base_url, post=False):
    """Fetches the physical device interface report and saves it to an Excel file."""
    inv_file = get_latest_inventory_file()
    if not inv_file:
        print("[ERROR] No inventory file found.")
        return

    df_inv = pd.read_excel(inv_file, engine="openpyxl")
    if post and "Action" in df_inv.columns:
        df_inv = df_inv[df_inv["Action"].astype(str).str.strip().str.upper() == "U"]

    results = []
    for index, row in df_inv.iterrows():
        site_id = row.get("site_id", "")
        mac = row.get("mac", "")
        name = row.get("name", "")
        if not site_id or not mac:
            continue

        stats_url = f"{base_url}/sites/{site_id}/stats/devices/00000000-0000-0000-1000-{mac}?type=gateway"
        response = requests.get(stats_url, headers={"Authorization": f"Token {token}"})
        log_api_request("GET", "retrieve_physical_device_interface_report", stats_url, None, response)

        if response.status_code == 200:
            data = response.json()
            if_stat = data.get("if_stat", {})
            for interface, stats in if_stat.items():
                results.append({
                    "Name": name,
                    "Site ID": site_id,
                    "MAC": mac,
                    "Interface": interface,
                    "Network Name": stats.get("network_name", ""),
                    "Status": "Up" if stats.get("up", False) else "Down"
                })
        else:
            results.append({"Name": name, "Site ID": site_id, "MAC": mac, "Interface": "Error", "Network Name": f"HTTP {response.status_code}", "Status": ""})

    df_results = pd.DataFrame(results)
    filename = "PhysicalDeviceInterfaceReport.xlsx" if not post else "Post-PhysicalDeviceInterfaceReport.xlsx"
    save_to_excel(df_results, filename=filename)

# ---------------------------
# Retrieve Full Device Interface Stats Report (Menu Item 5)
# ---------------------------
def retrieve_full_device_interface_stats_report(token, org_id, base_url, post=False):
    """Fetches the full device interface stats report and saves it to an Excel file."""
    inv_file = get_latest_inventory_file()
    if not inv_file:
        print("[ERROR] No inventory file found.")
        return

    df_inv = pd.read_excel(inv_file, engine="openpyxl")
    if post and "Action" in df_inv.columns:
        df_inv = df_inv[df_inv["Action"].astype(str).str.strip().str.upper() == "U"]

    all_data = []
    for _, row in df_inv.iterrows():
        site_id, mac, name = row.get("site_id", ""), row.get("mac", ""), row.get("name", "")
        if not site_id or not mac:
            continue

        url = f"{base_url}/sites/{site_id}/stats/ports/search?mac={mac}"
        response = requests.get(url, headers={"Authorization": f"Token {token}"})
        log_api_request("GET", "retrieve_full_device_interface_stats_report", url, None, response)

        if response.status_code == 200:
            data = response.json().get("results", [])
            for record in data:
                record["site_id"], record["name"] = site_id, name
            all_data.extend(data)
        else:
            print(f"[ERROR] Failed to retrieve data for site {site_id} and mac {mac}. HTTP {response.status_code}: {response.text}")

    df = pd.DataFrame(all_data)
    filename = "FullDeviceInterfaceStatsReport.xlsx" if not post else "Post-FullDeviceInterfaceStatsReport.xlsx"
    save_to_excel(df, filename=filename)
    
    
# ---------------------------
# Retrieve Device Peer Path Status Report (Menu Item 6)
# ---------------------------
def retrieve_device_peer_path_status_report(token, org_id, base_url, post=False):
    """
    Retrieves the device peer path status report.
    Saves the report to "DevicePeerPathReport.xlsx" or "Post-DevicePeerPathReport.xlsx" if post=True.
    """
    df_vpn = retrieve_vpn_peers_report(token, base_url, org_id, post=post)
    
    # Assign the correct filename
    filename = "DevicePeerPathReport.xlsx"
    if post:
        filename = "Post-" + filename
    
    # Save the DataFrame to Excel
    save_to_excel(df_vpn, filename=filename)
    print(f"[INFO] Device Peer Path Report saved as {filename}")

# ---------------------------
# Retrieve VPN Peers Report (Used by Peer Path Status Report)
# ---------------------------
def retrieve_vpn_peers_report(token, base_url, org_id, post=False):
    """
    Retrieves VPN Peers Report. If no data is found, returns an empty DataFrame with predefined columns.
    """
    latest_inv = get_latest_inventory_file()
    if not latest_inv:
        print("[ERROR] No inventory file found.")
        return pd.DataFrame()  # Return empty DataFrame instead of exiting
    
    df_inv = pd.read_excel(latest_inv, engine="openpyxl")
    
    # If 'post' is True, filter for devices marked for upgrade
    if post and "Action" in df_inv.columns:
        df_inv = df_inv[df_inv["Action"].astype(str).str.strip().str.upper() == "U"]
    
    unique_triples = df_inv[['site_id', 'mac', 'name']].dropna().drop_duplicates()
    all_records = []
    
    for _, row in unique_triples.iterrows():
        site = row['site_id']
        mac = row['mac']
        name = row['name']
        url = f"{base_url}/api/v1/orgs/{org_id}/stats/vpn_peers/search?mac={mac}"

        print(f"[DEBUG] Querying URL for site {site} and mac {mac}: {url}")

        headers = {"Authorization": f"Token {token}"}
        response = requests.get(url, headers=headers)
        log_api_request("GET", "retrieve_vpn_peers_report", url, None, response)

        if response.status_code == 200:
            data = response.json()
            if isinstance(data, dict) and "vpn_peers" in data:
                records = data["vpn_peers"]
            elif isinstance(data, list):
                records = data
            else:
                print(f"[WARNING] Unexpected data format for site {site} and mac {mac}.")
                continue
            
            for record in records:
                record.setdefault("site_id", site)
                record["name"] = name
            all_records.extend(records)
        else:
            print(f"[ERROR] Failed to retrieve data for site {site} and mac {mac}. HTTP {response.status_code}: {response.text}")
            log_error("HTTP error in VPN Peers Report", extra_info=f"Site {site} - HTTP {response.status_code}")

    if not all_records:
        print("[WARNING] No VPN peer data retrieved from any site. Returning an empty DataFrame.")
        return pd.DataFrame(columns=[
            "name", "site_id", "mac", "vlan_id", "latency", "vpn_role", "router_name", "type",
            "wan_name", "mos", "updated_from_event", "vpn_name", "loss", "peer_router_name",
            "peer_site_id", "peer_mac", "up", "is_active", "peer_port_id", "adjacent_address",
            "uptime", "network_interface", "mtu", "jitter", "hop_count", "port_id", "status",
            "last_seen", "org_id"
        ])  # Return an empty DataFrame with expected columns

    df = pd.DataFrame(all_records)

    # Select the relevant columns
    desired_fields = [
        "name", "site_id", "mac", "vlan_id", "latency", "vpn_role", "router_name", "type",
        "wan_name", "mos", "updated_from_event", "vpn_name", "loss", "peer_router_name",
        "peer_site_id", "peer_mac", "up", "is_active", "peer_port_id", "adjacent_address",
        "uptime", "network_interface", "mtu", "jitter", "hop_count", "port_id", "status",
        "last_seen", "org_id"
    ]
    cols = [field for field in desired_fields if field in df.columns]
    df = df[cols]
    df.sort_values(by="site_id", inplace=True)
    
    return df


# ---------------------------
# Access Point Status (Menu Item 7)
# ---------------------------
def access_point_status(token, org_id, base_url, allowed_sites=None, post=False):
    """
    Retrieves access point status from the Mist API for each site from the inventory,
    merges in the router name (from the inventory 'name' column), and outputs a formatted Excel report.
    
    If allowed_sites is provided, inventory is filtered to those site_ids.
    If post=True, it falls back to filtering by rows where "Action" == "U".
    """
    inventory_file = get_latest_inventory_file()
    if not inventory_file:
        print("[ERROR] No inventory file found.")
        log_error("No inventory file found in Access Point Status")
        return

    try:
        df_inventory = pd.read_excel(inventory_file, engine="openpyxl")
    except Exception as e:
        print(f"[ERROR] Could not read {inventory_file}: {e}")
        log_error("Error reading inventory in Access Point Status", extra_info=str(e))
        return

    if 'site_id' not in df_inventory.columns:
        print("[ERROR] The inventory file is missing the 'site_id' column.")
        log_error("Missing 'site_id' in Access Point Status")
        return

    # If filtering is needed for specific sites
    if allowed_sites is not None and len(allowed_sites) > 0:
        df_inventory = df_inventory[df_inventory["site_id"].isin(allowed_sites)]
    elif post and "Action" in df_inventory.columns:
        df_inventory = df_inventory[df_inventory["Action"].astype(str).str.strip().str.upper() == "U"]

    # Map site IDs to router names (if available)
    router_map = {}
    if 'name' in df_inventory.columns:
        router_map = df_inventory.groupby('site_id')['name'].last().to_dict()
    else:
        print("[WARNING] Inventory file missing 'name' column.")
        log_error("Missing 'name' column in Access Point Status")

    site_ids = df_inventory['site_id'].unique()
    print(f"[INFO] Found site IDs: {site_ids}")

    report_rows = []
    for site_id in site_ids:
        api_url = f"{base_url}/sites/{site_id}/stats/devices?type=ap"
        headers = {"Authorization": f"Token {token}", "Content-Type": "application/json"}

        try:
            response = requests.get(api_url, headers=headers)
            log_api_request("GET", "access_point_status", api_url, None, response)
        except Exception as e:
            print(f"[ERROR] Could not connect to {api_url} for site_id {site_id}: {e}")
            log_error("Connection error in Access Point Status", extra_info=f"site_id {site_id}: {e}")
            continue

        if response.status_code == 200:
            try:
                data = response.json()
            except Exception as e:
                print(f"[ERROR] Could not parse JSON for site_id {site_id}: {e}")
                log_error("JSON parse error in Access Point Status", extra_info=f"site_id {site_id}: {e}")
                continue

            if data:
                for device in data:
                    device_site_id = device.get("site_id", site_id)
                    status = device.get("status", "").strip() or "Missing"
                    report_rows.append({
                        "device_id": device.get("id", ""),
                        "name": device.get("name", ""),
                        "site_id": device_site_id,
                        "org_id": device.get("org_id", org_id if org_id else ""),
                        "status": status
                    })
            else:
                log_error("No devices found in Access Point Status", extra_info=f"site_id {site_id}")
                report_rows.append({
                    "device_id": "",
                    "name": "",
                    "site_id": site_id,
                    "org_id": org_id if org_id else "",
                    "status": "No devices found"
                })
        else:
            log_error("HTTP error in Access Point Status", extra_info=f"site_id {site_id} - HTTP {response.status_code}")
            report_rows.append({
                "device_id": "",
                "name": "",
                "site_id": site_id,
                "org_id": org_id if org_id else "",
                "status": f"HTTP {response.status_code}"
            })

    # Add router names to the report
    for row in report_rows:
        row["router_name"] = router_map.get(row["site_id"], "")

    if not report_rows:
        print("[INFO] No devices found for Access Point Status report.")
        log_error("No devices found in Access Point Status", extra_info="access_point_status")
        df = pd.DataFrame(columns=["router_name", "device_id", "name", "site_id", "org_id", "status"])
    else:
        df = pd.DataFrame(report_rows)

    if "router_name" not in df.columns:
        df["router_name"] = ""

    try:
        df.sort_values(by=["router_name"], inplace=True)
    except KeyError as e:
        print(f"[ERROR] Sorting error: {e}")
        log_error("Sorting error in Access Point Status", extra_info=str(e))
        return

    # Define the report filename
    filename = "AccessPointStats.xlsx"
    if post:
        filename = "Post-" + filename

    # Save the report
    save_to_excel(df, filename=filename)
    print(f"[INFO] Access Point Status report saved as {filename}")
    
# ---------------------------
# Post Reports (Menu Item 8)
# ---------------------------
def post_reports(token, org_id, base_url):
    """
    Generates post-upgrade reports by filtering devices marked with 'U' in the inventory.
    Runs all reporting functions and saves them with a 'Post-' prefix.
    """
    print("\n[INFO] Generating Post Reports for devices marked with 'U' in the inventory...")

    # Get latest inventory file
    inventory_file = get_latest_inventory_file()
    if not inventory_file:
        print("[ERROR] No inventory file found.")
        log_error("No inventory file found in Post Reports", extra_info="Menu Item 8")
        return

    try:
        df_inventory = pd.read_excel(inventory_file, engine="openpyxl")
    except Exception as e:
        print(f"[ERROR] Could not read {inventory_file}: {e}")
        log_error("Error reading inventory in Post Reports", extra_info=str(e))
        return

    # Ensure 'Action' column exists
    if "Action" not in df_inventory.columns:
        print("[ERROR] Inventory file missing 'Action' column.")
        log_error("Missing 'Action' column in Post Reports", extra_info="Menu Item 8")
        return

    # Filter devices marked for upgrade
    df_filtered = df_inventory[df_inventory["Action"].astype(str).str.strip().str.upper() == "U"]
    if df_filtered.empty:
        print("[INFO] No devices marked for upgrade ('U') found in inventory. No Post Reports generated.")
        log_error("No devices marked for upgrade in Post Reports", extra_info="Menu Item 8")
        return

    # Get allowed site IDs (for filtering reports)
    allowed_sites = df_filtered["site_id"].unique().tolist()
    print(f"[INFO] Allowed site IDs for Post Reports: {allowed_sites}")

    # Run all report functions with post=True
    retrieve_physical_device_interface_report(token, org_id, base_url, post=True)
    retrieve_full_device_interface_stats_report(token, org_id, base_url, post=True)
    retrieve_device_peer_path_status_report(token, org_id, base_url, post=True)
    access_point_status(token, org_id, base_url, allowed_sites=allowed_sites, post=True)
    retrieve_plan_valid_report(token, org_id, base_url, post=True)

    print("[INFO] All Post Reports generated successfully.")

# ---------------------------
# Retrieve Plan Valid Report (for Menu Item 8)
# ---------------------------
def retrieve_plan_valid_report(token, org_id, base_url, post=False):
    """
    Retrieves the plan valid report from the API using inventory data.
    Only runs for devices marked with 'U' (upgrade) when post=True.
    Saves report as 'Post-PlanValid.xlsx' if post=True, otherwise 'PlanValid.xlsx'.
    """
    print("\n[INFO] Retrieving Plan Valid Report...")

    # Get latest inventory file
    inventory_file = get_latest_inventory_file()
    if not inventory_file:
        print("[ERROR] No inventory file found.")
        return

    try:
        df_inventory = pd.read_excel(inventory_file, engine="openpyxl")
    except Exception as e:
        print(f"[ERROR] Could not read {inventory_file}: {e}")
        return

    # If running a post-upgrade report, filter devices marked for upgrade
    if post and "Action" in df_inventory.columns:
        df_inventory = df_inventory[df_inventory["Action"].astype(str).str.strip().str.upper() == "U"]

    if df_inventory.empty:
        print("[INFO] No devices available for Plan Valid Report.")
        return

    results = []

    for _, row in df_inventory.iterrows():
        site_id = row.get("site_id", "")
        mac = row.get("mac", "")
        name = row.get("name", "")

        if not site_id or not mac:
            continue

        # Construct API URL for fetching interface statistics
        api_url = f"{base_url}/sites/{site_id}/stats/devices/00000000-0000-0000-1000-{mac}?fields=if_stat"
        print(f"[DEBUG] Querying: {api_url}")

        try:
            response = requests.get(api_url, headers={"Authorization": f"Token {token}"})
            log_api_request("GET", "retrieve_plan_valid_report", api_url, None, response)
        except Exception as e:
            print(f"[ERROR] Failed to connect to {api_url}: {e}")
            continue

        if response.status_code == 200:
            try:
                data = response.json()
                if_stat = data.get("if_stat", {})

                if isinstance(if_stat, dict):
                    for interface, stats in if_stat.items():
                        network_name = stats.get("network_name", "Unknown")
                        up_status = "Up" if stats.get("up", False) else "Down"

                        results.append({
                            "Name": name,
                            "Site ID": site_id,
                            "MAC": mac,
                            "Interface": interface,
                            "Network Name": network_name,
                            "Status": up_status
                        })
                else:
                    results.append({
                        "Name": name,
                        "Site ID": site_id,
                        "MAC": mac,
                        "Interface": "N/A",
                        "Network Name": "N/A",
                        "Status": "No Data"
                    })

            except json.JSONDecodeError:
                print(f"[ERROR] JSON parsing failed for site {site_id} MAC {mac}.")
                continue
        else:
            print(f"[ERROR] Failed to retrieve data for site {site_id}. HTTP {response.status_code}: {response.text}")
            continue

    if not results:
        print("[INFO] No data retrieved for Plan Valid Report.")
        return

    # Convert to DataFrame
    df_results = pd.DataFrame(results, columns=["Name", "Site ID", "MAC", "Interface", "Network Name", "Status"])

    # Define filename
    filename = "PlanValid.xlsx"
    if post:
        filename = f"Post-{filename}"

    # Save results to Excel
    save_to_excel(df_results, filename=filename)
    print(f"[INFO] Plan Valid Report saved as {filename}.")



# ---------------------------
# Backup Org (Menu Item 9)
# ---------------------------
def create_backup_folder():
    """Creates a backup folder with a timestamp."""
    folder_name = f"backup-{datetime.now().strftime('%Y-%m-%d-%H-%M-%S')}"
    os.makedirs(folder_name, exist_ok=True)
    print(f"[INFO] Created backup folder: {folder_name}")
    return folder_name

def get_mist_data(org_id, token, endpoint, base_url):
    """Fetches data from Mist API for the specified endpoint."""
    url = f"{base_url}/orgs/{org_id}/{endpoint}"
    headers = {'Authorization': f"Token {token}"}
    
    response = requests.get(url, headers=headers)
    log_api_request("GET", "get_mist_data", url, None, response)

    if response.status_code != 200:
        print(f"[ERROR] Failed to fetch {endpoint}. HTTP {response.status_code}: {response.text}")
        log_error("Failed to fetch data in Backup Org", extra_info=f"{endpoint} - HTTP {response.status_code}")
        return []

    try:
        return response.json()
    except json.JSONDecodeError:
        print(f"[ERROR] Failed to parse JSON from {endpoint}")
        return []

def store_config_as_json(prefix, item, output_dir):
    """Saves a JSON object to a file in the backup directory."""
    filename = f"{prefix}-{item.get('id', 'unnamed')}.json"
    file_path = os.path.join(output_dir, filename)
    with open(file_path, 'w') as f:
        json.dump(item, f, indent=2)
    print(f"[INFO] Saved: {file_path}")

def backup_org(token, org_id, base_url):
    """Backs up configuration data from Mist API."""
    print("\n[INFO] Starting Backup Org process...")
    
    endpoints = {
        'apps': 'services',
        'networks': 'networks',
        'sites': 'sites',
        'hubs': 'deviceprofiles?type=gateway',
        'wan_edges': 'gatewaytemplates',
        'switches': 'networktemplates',
        'policies': 'servicepolicies'
    }

    backup_folder = create_backup_folder()
    
    for prefix, endpoint in endpoints.items():
        data = get_mist_data(org_id, token, endpoint, base_url)
        for item in data:
            store_config_as_json(prefix, item, backup_folder)

    print(f"[INFO] Backup completed. Files saved to {backup_folder}")
# ---------------------------
# Full Site Restore (Menu Item 10)
# ---------------------------
def read_json_files(prefix, directory="."):
    """Reads JSON files matching a prefix in the directory."""
    files = [f for f in os.listdir(directory) if f.startswith(prefix) and f.endswith('.json')]
    if not files:
        return []
    
    data_list = []
    for file_name in files:
        file_path = os.path.join(directory, file_name)
        with open(file_path, 'r') as f:
            try:
                data = json.load(f)
                if isinstance(data, dict):
                    data_list.append(data)
                elif isinstance(data, list):
                    data_list.extend(data)
            except json.JSONDecodeError:
                print(f"[ERROR] Could not parse JSON in {file_name}")
    return data_list

def submit_json(data_list, org_id, endpoint, token, base_url):
    """
    Submits JSON data to the Mist API for the specified endpoint.
    If a JSON file contains a list, each object is submitted separately.
    """
    url = f"{base_url}/orgs/{org_id}/{endpoint}"
    headers = {'Content-Type': 'application/json', 'Authorization': f"Token {token}"}

    if not data_list:
        print(f"[WARNING] No data to submit for {endpoint}.")
        log_error("No data to submit", extra_info=endpoint)
        return

    print(f"[INFO] Posting {len(data_list)} items to {endpoint}...")

    for item in data_list:
        if not isinstance(item, dict):
            print(f"[ERROR] Skipping invalid payload: Expected JSON object, got {type(item)}")
            log_error("Invalid payload in Full Site Restore", extra_info=str(type(item)))
            continue

        item = replace_org_id(item, org_id)
        post_url = url.split('?')[0]

        try:
            resp = requests.post(post_url, headers=headers, json=item)
            log_api_request("POST", "submit_json", post_url, item, resp)

            if resp.status_code in (200, 201):
                print(f"[INFO] Successfully posted to {endpoint}.")
                log_error("Successfully posted in Full Site Restore", extra_info=endpoint)

            elif resp.status_code == 409:  # Conflict - Might already exist
                print(f"[WARNING] Restore failed for {endpoint}: Resource may already exist.")
                log_error("Resource conflict in Full Site Restore", extra_info=f"Endpoint: {endpoint}, Response: {resp.text}")

            elif resp.status_code == 400 and "already exists" in resp.text.lower():
                print(f"[WARNING] Restore failed: {endpoint} - Item already exists.")
                log_error("Resource already exists in Full Site Restore", extra_info=f"Endpoint: {endpoint}, Response: {resp.text}")

            else:
                print(f"[ERROR] Restore failed for {endpoint}. HTTP {resp.status_code}: {resp.text}")
                log_error("Failed to post in Full Site Restore", extra_info=f"HTTP {resp.status_code}: {resp.text}")

        except Exception as e:
            print(f"[ERROR] Exception during POST to {endpoint}: {e}")
            log_error("Exception in Full Site Restore", extra_info=f"Endpoint: {endpoint}, Error: {str(e)}")


def full_site_restore(token, org_id, base_url):
    """
    Restores site configurations from a selected backup folder.

    - Allows restoring all components at once or selecting specific ones.
    - Ensures sites are restored first and their IDs are updated dynamically.
    - Implements retry logic to handle API downtime (HTTP 503).
    - Checks if an item already exists before attempting to restore.
    - Supports pressing 'X' to return to previous menus at every level.
    """

    # Step 1: List available backup folders
    backup_dirs = [d for d in os.listdir() if os.path.isdir(d) and d.startswith("backup-")]

    if not backup_dirs:
        print("[ERROR] No backup directories found.")
        return

    while True:
        print("\nAvailable backup folders:")
        for i, d in enumerate(backup_dirs, start=1):
            print(f"{i}. {d}")

        choice = input("\nSelect a backup folder by number (or press 'X' to return): ").strip().lower()
        if choice == 'x':
            print("[INFO] Returning to main menu.")
            return

        try:
            selected_folder = backup_dirs[int(choice) - 1]
            break
        except (ValueError, IndexError):
            print("[ERROR] Invalid selection. Please try again.")

    print(f"[INFO] Selected backup folder: {selected_folder}")

    # Step 2: Define all backup components
    components = {
        'apps': 'services',
        'networks': 'networks',
        'sites': 'sites',
        'hubs': 'deviceprofiles?type=gateway',
        'wan_edges': 'gatewaytemplates',
        'switches': 'networktemplates',
        'policies': 'servicepolicies'
    }

    # Step 3: Ask if restoring everything or only selected items
    while True:
        print("\nAvailable components to restore:")
        for i, (prefix, endpoint) in enumerate(components.items(), start=1):
            print(f" {i}. {prefix.capitalize()} ({endpoint})")
        print(" A. Restore All Components")
        print(" X. Return to Main Menu")

        restore_choice = input("\nSelect components to restore: ").strip().lower()

        if restore_choice == 'x':
            print("[INFO] Returning to main menu.")
            return
        elif restore_choice == 'a':
            selected_components = components  # Restore everything at once
            restore_all_json = True
            break
        else:
            try:
                selected_indices = [int(num.strip()) for num in restore_choice.split(",") if num.strip().isdigit()]
                selected_components = {list(components.keys())[i - 1]: list(components.values())[i - 1]
                                       for i in selected_indices if 1 <= i <= len(components)}
                restore_all_json = False
                if not selected_components:
                    raise ValueError
                break
            except (ValueError, IndexError):
                print("[ERROR] Invalid input. Please enter numbers separated by commas or 'A' for all.")

    print("\n[INFO] Starting restore process...")

    # Step 4: Restore Sites First (If Selected)
    new_site_id_mapping = {}
    if "sites" in selected_components:
        print("[INFO] Restoring Sites first...")
        site_data_list = read_json_files("sites", directory=selected_folder)

        for site_data in site_data_list:
            old_site_id = site_data.get("id")
            post_url = f"{base_url}/orgs/{org_id}/sites"
            headers = {'Content-Type': 'application/json', 'Authorization': f"Token {token}"}

            if check_if_exists(token, post_url, site_data):
                print(f"[INFO] Site {old_site_id} already exists. Skipping restore.")
                continue

            response = retry_request("POST", post_url, headers, json=replace_org_id(site_data, org_id))
            if response and response.status_code in (200, 201):
                response_data = response.json()
                new_site_id = response_data.get("id")
                if new_site_id:
                    new_site_id_mapping[old_site_id] = new_site_id
                    print(f"[INFO] Site created: old ID {old_site_id} -> new ID {new_site_id}")
                else:
                    print(f"[ERROR] Missing new site ID in response for site {old_site_id}")
            else:
                print(f"[ERROR] Failed to restore site {old_site_id}")

    # Step 5: Restore Other Selected Components
    if restore_all_json:
        # Restore all JSON files at once if 'Restore All' was selected
        for prefix, endpoint in selected_components.items():
            if prefix == "sites":
                continue

            print(f"\n[INFO] Restoring {prefix.capitalize()}...")
            data_list = read_json_files(prefix, directory=selected_folder)
            submit_json(data_list, org_id, endpoint, token, base_url)

    else:
        # Restore each component individually if selected manually
        for prefix, endpoint in selected_components.items():
            if prefix == "sites":
                continue

            while True:
                print(f"\n[INFO] Restoring {prefix.capitalize()}...")
                data_list = read_json_files(prefix, directory=selected_folder)

                if not data_list:
                    print(f"[WARNING] No data found for {prefix}. Skipping...")
                    break

                json_files = [f for f in os.listdir(selected_folder) if f.startswith(prefix) and f.endswith(".json")]
                print("\nAvailable JSON files to restore:")
                for i, json_file in enumerate(json_files, start=1):
                    print(f" {i}. {json_file}")
                print(" A. Restore All JSON files")
                print(" X. Return to Previous Menu")

                json_choice = input("\nSelect JSON files to restore: ").strip().lower()

                if json_choice == 'x':
                    print("[INFO] Returning to previous menu.")
                    break
                elif json_choice == 'a':
                    submit_json(data_list, org_id, endpoint, token, base_url)
                    break
                else:
                    try:
                        selected_json_indices = [int(num.strip()) for num in json_choice.split(",") if num.strip().isdigit()]
                        selected_json_files = [json_files[i - 1] for i in selected_json_indices if 1 <= i <= len(json_files)]
                        if not selected_json_files:
                            raise ValueError

                        selected_data_list = []
                        for json_file in selected_json_files:
                            with open(os.path.join(selected_folder, json_file), 'r') as f:
                                try:
                                    json_data = json.load(f)
                                    selected_data_list.append(json_data if isinstance(json_data, dict) else None)
                                except json.JSONDecodeError:
                                    print(f"[ERROR] Failed to parse {json_file}")

                        submit_json(selected_data_list, org_id, endpoint, token, base_url)
                        break

                    except (ValueError, IndexError):
                        print("[ERROR] Invalid input. Please enter numbers separated by commas or 'A' for all.")

    print("\n[INFO] Full site restore completed.")

# ---------------------------
# Utility Functions
# ---------------------------

def retry_request(method, url, headers, json=None, retries=3):
    """Retry API requests with exponential backoff for handling API downtime (503 errors)."""
    for attempt in range(retries):
        response = requests.request(method, url, headers=headers, json=json)
        if response.status_code != 503:
            return response
        print(f"[WARNING] API unavailable (503). Retrying in {2 ** attempt} seconds...")
        time.sleep(2 ** attempt)
    return response

def check_if_exists(token, url, data):
    """Check if an item already exists in Mist API before restoring."""
    headers = {'Authorization': f"Token {token}"}
    response = requests.get(url, headers=headers)
    return response.status_code == 200 and any(item["id"] == data.get("id") for item in response.json())
# ---------------------------
# Define Sections for Deletion
# ---------------------------
sections = {
    "1": {"name": "Sites", "list_path": "/orgs/{org_id}/sites", "delete_path": "/sites"},
    "2": {"name": "Applications", "list_path": "/orgs/{org_id}/services", "delete_path": "/orgs/{org_id}/services"},
    "3": {"name": "Networks", "list_path": "/orgs/{org_id}/networks", "delete_path": "/orgs/{org_id}/networks"},
    "4": {"name": "Hub Profiles", "list_path": "/orgs/{org_id}/deviceprofiles?type=gateway", "delete_path": "/orgs/{org_id}/deviceprofiles"},
    "5": {"name": "WAN Edges", "list_path": "/orgs/{org_id}/gatewaytemplates", "delete_path": "/orgs/{org_id}/gatewaytemplates"},
    "6": {"name": "Switches", "list_path": "/orgs/{org_id}/networktemplates", "delete_path": "/orgs/{org_id}/networktemplates"},
    "7": {"name": "Service Policies", "list_path": "/orgs/{org_id}/servicepolicies", "delete_path": "/orgs/{org_id}/servicepolicies"},
}

# ---------------------------
# List Items in a Section
# ---------------------------
def list_items(token, org_id, base_url, section):
    """Lists all items for the specified section from Mist API."""
    url = f"{base_url}{sections[section]['list_path'].format(org_id=org_id)}"
    headers = {"Authorization": f"Token {token}"}

    response = requests.get(url, headers=headers)

    if response.status_code == 200:
        items = response.json()
        if not items:
            print(f"[INFO] No {sections[section]['name']} found.")
            return []
        
        print(f"\n[INFO] Available {sections[section]['name']}:")
        for i, item in enumerate(items, start=1):
            print(f" {i}. {item.get('name', 'Unnamed')} (ID: {item['id']})")
        
        return items
    else:
        print(f"[ERROR] Failed to retrieve {sections[section]['name']}. HTTP {response.status_code}: {response.text}")
        return []

# ---------------------------
# Delete a Single Item
# ---------------------------
def delete_item(token, org_id, base_url, delete_path, item_id):
    """Deletes an item from Mist API by ID."""
    url = f"{base_url}{delete_path.format(org_id=org_id)}/{item_id}"
    headers = {"Authorization": f"Token {token}"}

    response = requests.delete(url, headers=headers)

    if response.status_code in [200, 204]:
        print(f"[INFO] Successfully deleted item with ID: {item_id}")
        return True
    else:
        print(f"[ERROR] Could not delete {item_id}. HTTP {response.status_code}")
        return False

# ---------------------------
# Remove All Items Across Sections
# ---------------------------
def remove_all_items(token, org_id, base_url):
    """Deletes all items across all sections with a strict double confirmation."""
    print("\n[WARNING] You are about to DELETE ALL ITEMS from all sections.")
    confirm1 = input("Are you ABSOLUTELY sure? Type 'yes' to confirm: ").strip().lower()

    if confirm1 not in ["yes", "delete all"]:
        print("[INFO] Operation canceled.")
        return

    confirm2 = input("Final confirmation: Type 'DELETE ALL' or 'yes' to proceed: ").strip().lower()

    if confirm2 not in ["delete all", "yes"]:
        print("[INFO] Operation aborted.")
        return

    deleted_items = []
    for section_key in sections.keys():
        section_name = sections[section_key]["name"]
        delete_path = sections[section_key]["delete_path"]
        items = list_items(token, org_id, base_url, section_key)

        if not items:
            continue

        print(f"[INFO] Deleting all {section_name}...")
        for item in items:
            if delete_item(token, org_id, base_url, delete_path, item['id']):
                deleted_items.append(f"{section_name}: {item.get('name', 'Unnamed')} (ID: {item['id']})")

    if deleted_items:
        print("\n[INFO] The following items were deleted:")
        for item in deleted_items:
            print(f" - {item}")
    else:
        print("[INFO] No items were deleted.")

    print("[INFO] ALL requested items have been removed successfully.")

# ---------------------------
# Interactive Menu for Deletions
# ---------------------------
def remove_org_items(token, org_id, base_url):
    """Interactive menu for removing org items from Mist API."""
    while True:
        print("\nChoose a section to manage for deletion:")
        for key in sorted(sections.keys()):
            print(f" {key}. {sections[key]['name']}")
        print(" 8. Remove All Items (DELETES EVERYTHING)")
        print(" 0. Return to Main Menu")
        
        section_choice = input("Enter your choice: ").strip()
        if section_choice == "0":
            print("[INFO] Returning to main menu.")
            return
        if section_choice == "8":
            remove_all_items(token, org_id, base_url)
            return
        if section_choice not in sections:
            print("[ERROR] Invalid choice.")
            continue

        section_name = sections[section_choice]["name"]
        delete_path = sections[section_choice]["delete_path"]
        
        items = list_items(token, org_id, base_url, section_choice)
        if not items:
            continue

        print("\nEnter the numbers of the items to delete, separated by commas (e.g., 1,3,5),")
        print("or type 'all' to delete everything in this section, or 'exit' to return.")

        user_input = input("Your choice: ").strip().lower()

        if user_input == "exit":
            print("[INFO] Returning to main menu.")
            return
        elif user_input == "all":
            confirm = input(f"Are you sure you want to delete ALL {section_name.lower()}? (yes/no): ").strip().lower()
            if confirm == "yes":
                for item in items:
                    delete_item(token, org_id, base_url, delete_path, item['id'])
            else:
                print("[INFO] Deletion canceled.")
        else:
            try:
                selections = [int(num.strip()) for num in user_input.split(",") if num.strip().isdigit()]
                for index in selections:
                    if 1 <= index <= len(items):
                        item = items[index - 1]
                        confirm = input(f"Delete {section_name} '{item.get('name', 'Unnamed')}' (ID: {item['id']})? (yes/no): ").strip().lower()
                        if confirm == "yes":
                            delete_item(token, org_id, base_url, delete_path, item['id'])
                        else:
                            print(f"[INFO] Skipped deletion of {section_name} '{item.get('name', 'Unnamed')}'.")
                    else:
                        print(f"[ERROR] Invalid selection: {index}")
            except ValueError:
                print("[ERROR] Invalid input. Please enter numbers separated by commas or 'all'.")

# ---------------------------
# Enhanced Logging Functions (dbug.xlsx)
# ---------------------------
def log_api_request(method, api_used, url, payload, response):
    """
    Logs API requests (GET, PUT, POST) to an Excel file 'dbug.xlsx'.
    Creates the file if it does not exist.
    Adds columns: TIME, FUNCTION CALLED, METHOD, API USED, PAYLOAD INFORMATION, DESCRIPTION OR ERROR INFO.
    """
    log_file = "dbug.xlsx"
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    stack = inspect.stack()
    caller_function = stack[1].function  # The function that called log_api_request()

    payload = json.dumps(payload, indent=2) if isinstance(payload, dict) else payload or ""

    response_info = f"HTTP {response.status_code}: {response.text}" if isinstance(response, requests.Response) else str(response)

    new_file = not os.path.exists(log_file)
    if new_file:
        wb = Workbook()
        ws = wb.active
        ws.append(["TIME", "FUNCTION CALLED", "METHOD", "API USED", "PAYLOAD INFORMATION", "DESCRIPTION OR ERROR INFO"])
    else:
        wb = load_workbook(log_file)
        ws = wb.active

    ws.append([timestamp, caller_function, method, url, payload, response_info])

    if new_file:
        header_fill = PatternFill(start_color="355E3B", end_color="355E3B", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:  # First row = headers
            cell.fill = header_fill
            cell.font = header_font

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2  

    wb.save(log_file)

# ---------------------------
# Non‑Blocking Key Press Helper
# ---------------------------
def wait_for_keypress(timeout):
    """Waits for a key press for a given timeout (in seconds)."""
    old_settings = termios.tcgetattr(sys.stdin)
    try:
        tty.setcbreak(sys.stdin.fileno())
        rlist, _, _ = select.select([sys.stdin], [], [], timeout)
        if rlist:
            return sys.stdin.read(1)
    finally:
        termios.tcsetattr(sys.stdin, termios.TCSADRAIN, old_settings)
    return None

# ---------------------------
# User Confirmation for Deletions
# ---------------------------
def confirm_action(message):
    """Prompts user for confirmation before proceeding."""
    confirmation = input(f"{message} (yes/no): ").strip().lower()
    return confirmation == "yes"

# ---------------------------
# API Request Wrapper with Error Handling
# ---------------------------
def api_request(method, url, token, payload=None, expected_status=200):
    """Handles API requests with retries and error logging."""
    headers = {"Authorization": f"Token {token}", "Content-Type": "application/json"}
    for attempt in range(3):  # Retry mechanism
        try:
            if method.upper() == "GET":
                response = requests.get(url, headers=headers)
            elif method.upper() == "POST":
                response = requests.post(url, headers=headers, json=payload)
            elif method.upper() == "PUT":
                response = requests.put(url, headers=headers, json=payload)
            elif method.upper() == "DELETE":
                response = requests.delete(url, headers=headers)
            else:
                raise ValueError(f"Unsupported HTTP method: {method}")

            log_api_request(method, "API Call", url, payload, response)

            if response.status_code == expected_status:
                return response.json()
            else:
                print(f"[ERROR] Attempt {attempt+1}: {method} request failed for {url}. HTTP {response.status_code}")
        except requests.exceptions.RequestException as e:
            print(f"[ERROR] Network error during {method} request to {url}: {e}")
    
    print(f"[ERROR] Failed {method} request after multiple attempts.")
    return None

# ---------------------------
# Main Execution
# ---------------------------
def main():
    """Main execution loop for the program."""
    token, org_id, base_url = read_token_org_url("Token-Org-URL.txt")
    while True:
        display_menu()
        choice = input("Choose an option: ").strip()
        
        if choice == "1":
            devices = get_gateway_inventory(org_id, token, base_url)
            if devices:
                save_to_excel(pd.DataFrame(devices), filename="inventoryStats.xlsx")
        
        elif choice == "2":
            latest_file = get_latest_inventory_file()
            if latest_file:
                process_inventory_actions(latest_file, token, base_url)
            else:
                print("[ERROR] No inventory file found.")
        
        elif choice == "3":
            latest_file = get_latest_inventory_file()
            if latest_file:
                check_upgrade_status(latest_file, token, base_url)
            else:
                print("[ERROR] No inventory file found.")
        
        elif choice == "4":
            retrieve_physical_device_interface_report(token, org_id, base_url, post=False)
        
        elif choice == "5":
            retrieve_full_device_interface_stats_report(token, org_id, base_url, post=False)
        
        elif choice == "6":
            retrieve_device_peer_path_status_report(token, org_id, base_url, post=False)
        
        elif choice == "7":
            access_point_status(token, org_id, base_url)
        
        elif choice == "8":
            post_reports(token, org_id, base_url)
        
        elif choice == "9":
            backup_org(token, org_id, base_url)
        
        elif choice == "10":
            full_site_restore(token, org_id, base_url)
        
        elif choice == "11":
            remove_org_items(token, org_id, base_url)
        
        elif choice == "0":
            print("[INFO] Exiting program.")
            sys.exit()
        
        else:
            print("[ERROR] Invalid choice. Please select again.")

if __name__ == "__main__":
    main()

